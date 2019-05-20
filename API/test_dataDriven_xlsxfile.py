import requests
import jsonpath
import json
import openpyxl

def test_addMultipleStudent():
    API_URL = "http://thetestingworldapi.com/api/studentsDetails"
    file = open('D:\\Ankit\\apiwork\\addstudent.json', 'r')
    body = json.loads(file.read())
    vk = openpyxl.load_workbook("D:\\Ankit\\apiwork\\testdata.xlsx")
    sh = vk['Sheet1']
    rows = sh.max_row
    for i in range(2,rows+1):
        first_name = sh.cell(row=i, column=1)
        mid_name = sh.cell(row=i, column=2)
        last_name = sh.cell(row=i, column=3)
        dob = sh.cell(row=i, column=4)
        body['first_name'] = first_name.value
        body['middle_name'] = mid_name.value
        body['last_name'] = last_name.value
        body['date_of_birth'] = dob.value
        response = requests.post(API_URL, body)
        st_id = jsonpath.jsonpath(response.json(), 'id')
        print(st_id[0])
