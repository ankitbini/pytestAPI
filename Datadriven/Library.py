import json
import jsonpath
import requests
import openpyxl

class Common:

    def __init__(self, filepath, sheet_name):
        global vk
        global sh
        vk = openpyxl.load_workbook(filepath)
        sh = vk[sheet_name]

    def fetchRowCount(self):
        rows = sh.max_row
        return rows


    def fetchColumnCount(self):
        col = sh.max_column
        return col

    def fetchKeyName(self):
        c = sh.max_column
        ls = []
        for i in range(1, c+1):
            cell = sh.cell(row=1, column=i)
            ls.insert(i-1, cell.value)
        return ls

    def updateRequestWithData(self, row_number, json_request, key_list):
        c = sh.max_column
        for i in range(1, c+1):
            cell = sh.cell(row=row_number, column=i)
            json_request[key_list[i-1]] = cell.value

        return json_request
